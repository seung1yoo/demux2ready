
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class MetileneData:

    def __init__(self):
        pass

    def write_sheet(self, df, ws, ws_title):
        """
        DataFrame을 주어진 시트에 쓰고 스타일을 적용하는 함수
        """
        ws.title = ws_title

        # DataFrame을 시트로 쓰기
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 헤더 스타일 적용
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        for cell in ws[1]:  # 첫 번째 행(헤더)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 색상 설정
        green_fill = PatternFill("solid", fgColor="C6EFCE")
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        blue_fill = PatternFill("solid", fgColor="DDEBF7")
        orange_fill = PatternFill("solid", fgColor="FDE9D9")

        # 
        _cols = [col for col in ws[1] if col.value.endswith("q-value")]
        for col in _cols:
            for cell in ws.iter_rows(min_row=2, min_col=col.col_idx, max_col=col.col_idx, max_row=ws.max_row):
                cell = cell[0]
                if float(cell.value) < 0.05:
                    cell.fill = blue_fill
                else:
                    cell.fill = red_fill
        # 
        _cols = [col for col in ws[1] if col.value.endswith("mean_methylation_delta")]
        for col in _cols:
            for cell in ws.iter_rows(min_row=2, min_col=col.col_idx, max_col=col.col_idx, max_row=ws.max_row):
                cell = cell[0]
                if abs(cell.value) < 20.0:
                    cell.fill = red_fill
                elif abs(cell.value) < 80.0:
                    cell.fill = orange_fill
                else:
                    cell.fill = green_fill

        # 각 열의 너비 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 컬럼 문자 가져오기 (A, B, C 등)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width



    def write_excel(self, cpg_file, chg_file, chh_file, ofn):
        """
        여러 TSV 파일을 하나의 엑셀 파일의 여러 시트에 저장하는 함수
        """
        # 엑셀 워크북 생성
        wb = Workbook()

        # 각각의 파일을 DataFrame으로 읽기
        cpg_df = pd.read_csv(cpg_file, sep="\t")
        chg_df = pd.read_csv(chg_file, sep="\t")
        chh_df = pd.read_csv(chh_file, sep="\t")

        # 각 시트에 데이터를 쓰고 스타일 적용
        ws_cpg = wb.active
        self.write_sheet(cpg_df, ws_cpg, 'CpG')

        ws_chg = wb.create_sheet(title="CHG")
        self.write_sheet(chg_df, ws_chg, 'CHG')

        ws_chh = wb.create_sheet(title="CHH")
        self.write_sheet(chh_df, ws_chh, 'CHH')

        # 엑셀 파일 저장
        wb.save(ofn)


def main(args):
    metilene = MetileneData()
    metilene.write_excel(args.cpg_result, args.chg_result, args.chh_result, args.outfn)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--cpg-result', default="DMR001/CG/metilene_Donor_P05.output.anno")
    parser.add_argument('--chg-result', default="DMR001/CHG/metilene_Donor_P05.output.anno")
    parser.add_argument('--chh-result', default="DMR001/CHH/metilene_Donor_P05.output.anno")
    parser.add_argument('--outfn', default="DMR001_Donor_P05.xlsx")
    args = parser.parse_args()
    main(args)