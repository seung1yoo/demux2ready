
from ossaudiodev import control_names
from re import I
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from pathlib import Path
import logging
logging.basicConfig(level=logging.DEBUG)
import sys
import subprocess

class MetilenePrepare:
    bismark_split_cx_by_context_sh = Path('/storage2/User/siyoo/module/demux2ready/lib/bismark_split_cx_by_context.sh')
    bismark_cx_to_bedGraph_sh = Path('/storage2/User/siyoo/module/demux2ready/lib/bismark_cx_to_bedGraph.sh')
    
    def mkcmd_split_cx_by_context_sh(self, sample, is_dedup):
        cmd = list()
        cmd.append(str(self.bismark_split_cx_by_context_sh))
        cmd.append('demux2ready/bismark')
        cmd.append(sample)
        cmd.append(f'demux2ready/bismark/{sample}_split_cx_by_context')
        cmd.append(f'"{is_dedup}"')
        return cmd
    
    def mkcmd_cx_to_bedGraph_sh(self, sample, context, min_depth, is_dedup):
        cmd = list()
        cmd.append(str(self.bismark_cx_to_bedGraph_sh))
        cmd.append(f'demux2ready/bismark/{sample}_split_cx_by_context')
        cmd.append(sample)
        cmd.append(context)
        cmd.append(min_depth)
        cmd.append(f'"{is_dedup}"')
        return cmd


class MetileneData:

    bedGraph_path_dic = dict()
    metilene_path_dic = dict()

    def load_cx_to_bedGraph_sh_results(self, sample, context, is_dedup):
        self.bedGraph_path_dic.setdefault(sample, {})
        _path = Path('demux2ready/bismark') / f"{sample}_split_cx_by_context" / f"{sample}_R1_bismark_bt2_pe.{is_dedup}bedGraph.{context}.gz"
        if _path.exists():
            self.bedGraph_path_dic[sample][context] = _path
            logging.info(f"Load bedGraph path : {str(_path)}")
        else:
            logging.info(f"Failed to load bedGraph path : {str(_path)} is not exists.")
            sys.exit(1)
    
    def load_metilene_results(self, dmr_id, control_name, case_name, context):
        self.metilene_path_dic.setdefault(dmr_id, {})
        _path = self.outdir / dmr_id / context / f"{self.outprefix}_{control_name}_{case_name}.output.anno"
        if _path.exists():
            self.metilene_path_dic[dmr_id][context] = _path
            logging.info(f"Load metilene path : {str(_path)}")
        else:
            logging.info(f"Failed to load metilene path : {str(_path)} is not exists.")

    def write_sheet(self, df, ws, ws_title):
        ws.title = ws_title

        # DataFrame을 시트로 쓰기
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 헤더 스타일 적용
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        for cell in ws[1]:  # 첫 번째 행(헤더)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 색상 설정
        green_fill = PatternFill("solid", fgColor="C6EFCE")
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        blue_fill = PatternFill("solid", fgColor="DDEBF7")
        orange_fill = PatternFill("solid", fgColor="FDE9D9")

        # 
        _cols = [col for col in ws[1] if col.value.endswith("q-value")]
        for col in _cols:
            for cell in ws.iter_rows(min_row=2, min_col=col.col_idx, max_col=col.col_idx, max_row=ws.max_row):
                cell = cell[0]
                if float(cell.value) < 0.05:
                    cell.fill = blue_fill
                else:
                    cell.fill = red_fill
        # 
        _cols = [col for col in ws[1] if col.value.endswith("mean_methylation_delta")]
        for col in _cols:
            for cell in ws.iter_rows(min_row=2, min_col=col.col_idx, max_col=col.col_idx, max_row=ws.max_row):
                cell = cell[0]
                if abs(cell.value) < 20.0:
                    cell.fill = red_fill
                elif abs(cell.value) < 80.0:
                    cell.fill = orange_fill
                else:
                    cell.fill = green_fill

        # 각 열의 너비 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 컬럼 문자 가져오기 (A, B, C 등)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width



    def write_excel(self, dmr_id, cpg_file, chg_file, chh_file):
        wb = Workbook()

        cpg_df = pd.read_csv(cpg_file, sep="\t")
        chg_df = pd.read_csv(chg_file, sep="\t")
        chh_df = pd.read_csv(chh_file, sep="\t")

        ws_cpg = wb.active
        self.write_sheet(cpg_df, ws_cpg, 'CpG')

        ws_chg = wb.create_sheet(title="CHG")
        self.write_sheet(chg_df, ws_chg, 'CHG')

        ws_chh = wb.create_sheet(title="CHH")
        self.write_sheet(chh_df, ws_chh, 'CHH')

        wb.save(str( self.outdir / dmr_id / f"{dmr_id}.Result_table.xlsx"))


class MetileneExecute:

    metilene_exe = Path("metilene") # in conda env
    metilene_input_pl = Path("metilene_input.pl") # in conda env
    metilene_output_pl = Path("metilene_output.pl") # in conda env

    def mkcmd_metilene_input_pl(self, dmr_id, control_name, controls, case_name, cases, context):
        a_outdir = self.outdir / dmr_id / context
        a_outdir.mkdir(exist_ok=True, parents=True)
        a_result = a_outdir / f"{self.outprefix}_{control_name}_{case_name}.input"

        cmd = [str(self.metilene_input_pl)]
        cmd.append('--in1')
        cmd.append(','.join([str(self.bedGraph_path_dic[x][context]) for x in controls]))
        cmd.append('--in2')
        cmd.append(','.join([str(self.bedGraph_path_dic[x][context]) for x in cases]))
        cmd.append('--h1')
        cmd.append(control_name)
        cmd.append('--h2')
        cmd.append(case_name)
        cmd.append('--out')
        cmd.append(str(a_result))
        return cmd
    
    def mkcmd_metilene_input_sort(self, dmr_id, control_name, controls, case_name, cases, context):
        a_outdir = self.outdir / dmr_id / context
        a_outdir.mkdir(exist_ok=True, parents=True)
        a_result = a_outdir / f"{self.outprefix}_{control_name}_{case_name}.input.sort"

        cmd = ['cat']
        cmd.append(str(a_outdir / f"{self.outprefix}_{control_name}_{case_name}.input"))
        cmd.append('|')
        cmd.append('head -n 1 && tail -n +2')
        cmd.append('|')
        cmd.append('sort -k1,1 -k2,2n')
        cmd.append('|')
        cmd.append('uniq')
        cmd.append('>')
        cmd.append(str(a_result))

        return cmd
    
    def execute_metilene(self, dmr_id, control_name, case_name, context):
        a_outdir = self.outdir / dmr_id / context
        a_outdir.mkdir(exist_ok=True, parents=True)
        a_result = a_outdir / f"{self.outprefix}_{control_name}_{case_name}.output"

        cmd = [str(self.metilene_exe)]
        cmd.append('-t')
        cmd.append('4')
        cmd.append('-a')
        cmd.append(control_name)
        cmd.append('-b')
        cmd.append(case_name)
        cmd.append(str(a_outdir / f"{self.outprefix}_{control_name}_{case_name}.input.sort"))
        cmd.append('|')
        cmd.append('awk')
        cmd.append("'$2 < $3 {print $0}'")
        cmd.append('|')
        cmd.append('sort')
        cmd.append('-V')
        cmd.append('-k1,1')
        cmd.append('-k2,2n')
        cmd.append('>')
        cmd.append(str(a_result))
        return cmd

    def mkcmd_metilene_output_pl(self, dmr_id, control_name, case_name, context):
        a_outdir = self.outdir / dmr_id / context
        a_outdir.mkdir(exist_ok=True, parents=True)
        a_result = a_outdir / f"{self.outprefix}_{control_name}_{case_name}.output.filter_qval.0.05.out"

        cmd = [str(self.metilene_output_pl)]
        cmd.append('-q')
        cmd.append(str(a_outdir / f"{self.outprefix}_{control_name}_{case_name}.output"))
        cmd.append('-o')
        cmd.append(str(a_outdir / f"{self.outprefix}_{control_name}_{case_name}.output.filter"))
        cmd.append('-a')
        cmd.append(control_name)
        cmd.append('-b')
        cmd.append(case_name)
        return cmd

    def mkcmd_metilene_anno(self, dmr_id, control_name, case_name, context, anno_bed):
        a_outdir = self.outdir / dmr_id / context
        a_outdir.mkdir(exist_ok=True, parents=True)
        a_result = a_outdir / f"{self.outprefix}_{control_name}_{case_name}.output.anno"

        cmd = ['bedtools']
        cmd.append('intersect')
        cmd.append('-a')
        cmd.append(str(a_outdir / f"{self.outprefix}_{control_name}_{case_name}.output"))
        cmd.append('-b')
        cmd.append(anno_bed)
        cmd.append('-wa')
        cmd.append('-loj')
        cmd.append('|')
        cmd.append('sed')
        cmd.append("\"1ichr\\tstart\\tstop\\tq-value\\tmean_methylation_delta\\tcpg_count\\tp-MWU\\tp-2D_KS\\t"
                   f"mean_{control_name}\\tmean_{case_name}\\t"
                   "anno_chr\\tanno_start\\tanno_stop\\tanno_gene\\tanno_score\\tanno_strand\"")
        cmd.append("|")
        cmd.append("(head -n 1 && tail -n +2 | sort -g -k4,4 -k5,5n)")
        cmd.append("|")
        cmd.append("awk 'BEGIN {OFS=\"\\t\"} NR==1 {print; next} $11 != \".\" {print}"
                   " $11 == \".\" {$11=\"\"; for(i=12; i<=NF; i++) $i=\"\"; print}'")
        cmd.append('>')
        cmd.append(str(a_result))
        return cmd


class Metilene(MetilenePrepare, MetileneData, MetileneExecute):
    def __init__(self, samples, outdir, outprefix):
        self.samples = samples
        self.outdir = Path(outdir)
        self.outdir.mkdir(exist_ok=True, parents=True)
        self.outprefix = outprefix
        self.dmr_comp_dic = dict()
    
    def parse_samples(self, _str):
        return _str.split(',')

    def load_dmr_comp(self, dmr_comp):
        for dmr_set in dmr_comp:
            dmr_id = dmr_set[0]
            control_name = dmr_set[1]
            controls = self.parse_samples(dmr_set[2])
            case_name = dmr_set[3]
            cases = self.parse_samples(dmr_set[4])
            self.dmr_comp_dic.setdefault(dmr_id, {}).setdefault('control_name', control_name)
            self.dmr_comp_dic.setdefault(dmr_id, {}).setdefault('controls', controls)
            self.dmr_comp_dic.setdefault(dmr_id, {}).setdefault('case_name', case_name)
            self.dmr_comp_dic.setdefault(dmr_id, {}).setdefault('cases', cases)
            logging.info(f"DMR SET : {dmr_id}")
            logging.info(f"{control_name} : {controls}")
            logging.info(f"{case_name} : {cases}")
    
    def do_prepare(self, min_depth, is_dedup, is_run_cmd=True):
        for sample in self.samples:
            cmd = self.mkcmd_split_cx_by_context_sh(sample, is_dedup)
            if is_run_cmd:
                logging.debug(' '.join(cmd))
                stdout, stderr = run_cmd(cmd)
                logging.info(f"stdout : {stdout}")
                logging.info(f"stderr : {stderr}")
            for context in ['CG', 'CHG','CHH']:
                cmd = self.mkcmd_cx_to_bedGraph_sh(sample, context, min_depth, is_dedup)
                if is_run_cmd:
                    logging.debug(' '.join(cmd))
                    stdout, stderr = run_cmd(cmd)
                    logging.info(f"stdout : {stdout}")
                    logging.info(f"stderr : {stderr}")
                self.load_cx_to_bedGraph_sh_results(sample, context, is_dedup)
        
    def do_execute(self, dmr_comp, anno_bed, is_run_cmd=True):
        self.load_dmr_comp(dmr_comp)
        for dmr_id, sample_dic in self.dmr_comp_dic.items():
            control_name = sample_dic['control_name']
            controls = sample_dic['controls']
            case_name = sample_dic['case_name']
            cases = sample_dic['cases']
            for context in ['CG', 'CHG','CHH']:
                cmd = self.mkcmd_metilene_input_pl(dmr_id, control_name, controls, case_name, cases, context)
                if is_run_cmd:
                    logging.debug(' '.join(cmd))
                    stdout, stderr = run_cmd(cmd)
                    logging.info(f"stdout : {stdout}")
                    logging.info(f"stderr : {stderr}")
                cmd = self.mkcmd_metilene_input_sort(dmr_id, control_name, controls, case_name, cases, context)
                if is_run_cmd:
                    logging.debug(' '.join(cmd))
                    stdout, stderr = run_cmd(cmd, shell=True)
                    logging.info(f"stdout : {stdout}")
                    logging.info(f"stderr : {stderr}")
                cmd = self.execute_metilene(dmr_id, control_name, case_name, context)
                if is_run_cmd:
                    logging.debug(' '.join(cmd))
                    stdout, stderr = run_cmd(cmd, shell=True)
                    logging.info(f"stdout : {stdout}")
                    logging.info(f"stderr : {stderr}")
                cmd = self.mkcmd_metilene_output_pl(dmr_id, control_name, case_name, context)
                if is_run_cmd:
                    logging.debug(' '.join(cmd))
                    stdout, stderr = run_cmd(cmd)
                    logging.info(f"stdout : {stdout}")
                    logging.info(f"stderr : {stderr}")
                cmd = self.mkcmd_metilene_anno(dmr_id, control_name, case_name, context, anno_bed)
                if is_run_cmd:
                    logging.debug(' '.join(cmd))
                    stdout, stderr = run_cmd(cmd, shell=True)
                    logging.info(f"stdout : {stdout}")
                    logging.info(f"stderr : {stderr}")
                self.load_metilene_results(dmr_id, control_name, case_name, context)
    
    def do_summary(self):
        for dmr_id, context_dic in self.metilene_path_dic.items():
            cpg_result = context_dic['CG']
            chg_result = context_dic['CHG']
            chh_result = context_dic['CHH']
            self.write_excel(dmr_id, cpg_result, chg_result, chh_result)


def run_cmd(cmd, shell=False):

    if shell:
        try:
            result = subprocess.run(' '.join(cmd), shell=True, text=True, capture_output=True)
            if result.returncode != 0:
                logging.error(f"Error executing: {' '.join(cmd)}\nError : {result.stderr}")
                sys.exit(1)
            return result.stdout, result.stderr
        except Exception as e:
            logging.error(f"Exception while executing: {' '.join(cmd)}\n{e}")
            sys.exit(1)
    else:
        try:
            result = subprocess.run(cmd, text=True, capture_output=True)
            if result.returncode != 0:
                logging.error(f"Error executing: {' '.join(cmd)}\nError : {result.stderr}")
                sys.exit(1)
            return result.stdout, result.stderr
        except Exception as e:
            logging.error(f"Exception while executing: {' '.join(cmd)}\n{e}")
            sys.exit(1)



def main(args):

    metilene = Metilene(args.samples, args.outdir, args.outprefix)

    if args.is_run_prepare:
        metilene.do_prepare(args.min_depth, args.is_dedup, is_run_cmd=True)
    else:
        metilene.do_prepare(args.min_depth, args.is_dedup, is_run_cmd=False)
    
    if args.is_run_execute:
        metilene.do_execute(args.dmr_comp, args.anno_bed, is_run_cmd=True)
    else:
        metilene.do_execute(args.dmr_comp, args.anno_bed, is_run_cmd=False)

    metilene.do_summary()




if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--samples', nargs="+", default=["YiP3P15_1","YiP3P15_2","YiP3P15_3",
                                                         "A7_1", "A7_2", "A7_3",
                                                         "B2_1", "B2_2", "B2_3"])
    parser.add_argument('--min-depth', default='5', type=str)
    parser.add_argument('--dmr-comp', action='append', nargs=5,
                        metavar=('dmr_id', 'control_group_name', 'control_samples', 'case_group_name', 'case_samples'))
    parser.add_argument('--outdir', default="demux2ready/metilene")
    parser.add_argument('--outprefix', default="metilene")
    parser.add_argument('--anno-bed', default="/storage2/User/siyoo/module/demux2ready/src/MANE.GRCh38.v1.3.summary.chr.sorted.bed")
    parser.add_argument('--is-dedup', default="deduplicated.")
    parser.add_argument('--is-run-prepare', action='store_true')
    parser.add_argument('--is-run-execute', action='store_true')
    args = parser.parse_args()
    main(args)

