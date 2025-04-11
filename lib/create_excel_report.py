#!/usr/bin/env python3
import argparse
import logging
import pandas as pd
from pathlib import Path
from typing import Dict, Any
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, numbers, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def parse_args() -> argparse.Namespace:
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='Create Excel report from sequencing data')
    parser.add_argument(
        '--md5sum',
        type=str,
        default='demux2ready/md5sum.summary.tsv',
        help='Path to md5sum summary file'
    )
    parser.add_argument(
        '--fastp',
        type=str,
        default='demux2ready/fastp.summary.tsv',
        help='Path to fastp summary file'
    )
    parser.add_argument(
        '--s3upload',
        type=str,
        default='atgcu-util.s3-upload.result.tsv',
        help='Path to S3 upload result file'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='PROJECT_ID-RawFastq-Info-vDATE.xlsx',
        help='Output Excel file name'
    )
    return parser.parse_args()

def apply_table_style(worksheet, df):
    """Apply consistent table styling to the worksheet"""
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Apply styles to header row
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Apply styles to data rows
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border
            
            # Left align specific columns
            column_name = df.columns[col-1]
            if column_name in ['File Name', 'S3 URL']:
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment

def adjust_column_widths(worksheet, df):
    """Adjust column widths based on content"""
    for idx, col in enumerate(df.columns, 1):
        # Special handling for Presigned URL column
        if col == 'Presigned URL':
            worksheet.column_dimensions[get_column_letter(idx)].width = len('Presigned URL') + 2
            continue
            
        # For other columns, adjust based on content
        max_length = max(
            df[col].astype(str).apply(len).max(),
            len(str(col))
        )
        # Add some padding
        adjusted_width = max_length + 2
        # Set the column width
        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

def format_column_as_number(worksheet, df, column_name):
    """Format File Size column as numbers with thousand separators"""
    # Find the File Size column index
    size_col_idx = None
    for idx, col in enumerate(df.columns, 1):
        if col in [column_name]:
            size_col_idx = idx
            break
    
    if size_col_idx is None:
        return
    
    # Format each cell in the File Size column
    for row in range(2, len(df) + 2):  # Start from row 2 (after header)
        cell = worksheet.cell(row=row, column=size_col_idx)
        if cell.value:
            # Convert to number format with thousand separators
            cell.number_format = '#,##0'

def add_hyperlinks(worksheet, df):
    """Add hyperlinks to the Presigned URL column"""
    # Find the Presigned URL column index
    url_col_idx = None
    for idx, col in enumerate(df.columns, 1):
        if col == 'Presigned URL':
            url_col_idx = idx
            break
    
    if url_col_idx is None:
        return
    
    # Add hyperlinks to each cell in the Presigned URL column
    for row in range(2, len(df) + 2):  # Start from row 2 (after header)
        cell = worksheet.cell(row=row, column=url_col_idx)
        if cell.value and isinstance(cell.value, str):
            # Store the URL for the hyperlink
            url = cell.value
            # Change the display text to "Download"
            cell.value = "Download"
            # Add the hyperlink
            cell.hyperlink = url
            # Style the cell
            cell.font = Font(color="0000FF", underline="single")
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


class DataLoader:

    def create_md5sum_info(self, md5sum_path: Path) -> pd.DataFrame:
        md5sum_df = pd.read_csv(md5sum_path, sep='\s+', header=None, names=['md5', 'file_path'])
        md5sum_df['file_name'] = md5sum_df['file_path'].apply(lambda x: Path(x).name)

        # Select and rename columns
        result_df = pd.DataFrame()
        result_df['File Name'] = md5sum_df['file_name']
        result_df['MD5 Checksum'] = md5sum_df['md5']

        return result_df

    def create_download_info(self, s3upload_path: Path) -> pd.DataFrame:
        # Read S3 upload result
        s3upload_df = pd.read_csv(s3upload_path, sep='\t')
        
        # Convert File_size(Bytes) to numeric, removing commas
        s3upload_df['File_size(Bytes)'] = s3upload_df['File_size(Bytes)'].str.replace(',', '').astype(float)
        
        # Select and rename columns
        result_df = pd.DataFrame()
        result_df['File Name'] = s3upload_df['File_name']
        result_df['File Size (Bytes)'] = s3upload_df['File_size(Bytes)']
        result_df['S3 URL'] = s3upload_df['S3_url']
        result_df['Presigned URL'] = s3upload_df['Presigned_url']
        result_df['Create Date'] = s3upload_df['Create_date']
        result_df['Expiry Date'] = s3upload_df['Expiry_date']
        
        return result_df

    def create_download_info_with_md5sum(self, s3upload_path: Path, md5sum_path: Path) -> pd.DataFrame:
        # Read md5sum file
        md5sum_df = pd.read_csv(md5sum_path, sep='\s+', header=None, names=['md5', 'file_path'])
        md5sum_df['file_name'] = md5sum_df['file_path'].apply(lambda x: Path(x).name)
        
        # Read S3 upload result
        s3upload_df = pd.read_csv(s3upload_path, sep='\t')
        
        # Convert File_size(Bytes) to numeric, removing commas
        s3upload_df['File_size(Bytes)'] = s3upload_df['File_size(Bytes)'].str.replace(',', '').astype(float)
        
        # Merge dataframes with outer join to include all files
        download_info = pd.merge(
            s3upload_df,
            md5sum_df[['file_name', 'md5']],
            left_on='File_name',
            right_on='file_name',
            how='outer'
        )
        
        # Select and rename columns
        result_df = pd.DataFrame()
        result_df['File Name'] = download_info['File_name']
        result_df['MD5 Checksum'] = download_info['md5']
        result_df['File Size (Bytes)'] = download_info['File_size(Bytes)']
        result_df['S3 URL'] = download_info['S3_url']
        result_df['Presigned URL'] = download_info['Presigned_url']
        result_df['Create Date'] = download_info['Create_date']
        result_df['Expiry Date'] = download_info['Expiry_date']
        
        return result_df

    def create_sequencing_info(self, fastp_path: Path) -> pd.DataFrame:
        """Create sequencing information table"""
        # Read fastp summary
        sequencing_info = pd.read_csv(fastp_path, sep='\t')
        
        # Rename columns for better readability
        column_mapping = {
            'Samples': 'Sample ID',
            'ReadsCount': 'Total Reads',
            'BasesCount': 'Total Bases',
            'BasesCountGb': 'Total Bases (GB)',
            'GCRate': 'GC Rate (%)',
            'Q20BaseRate': 'Q20 Rate (%)',
            'Q30BaseRate': 'Q30 Rate (%)',
            'DupRate': 'Duplication Rate (%)',
            'InsertSize': 'Insert Size (bp)',
            'R1MeanLen': 'R1 Mean Length (bp)',
            'R2MeanLen': 'R2 Mean Length (bp)',
            'GoodReadRate': 'Good Read Rate (%)'
        }
        
        sequencing_info = sequencing_info.rename(columns=column_mapping)
        return sequencing_info

class ReportGenerator(DataLoader):

    def __init__(self, args: argparse.Namespace):
        self.fastp_path = Path(args.fastp)
        self.md5sum_path = Path(args.md5sum)
        self.s3upload_path = Path(args.s3upload)
        self.output_path = Path(args.output)

        self.report_type = self.define_report_type()

    def define_report_type(self) -> str:
        if self.fastp_path.exists() and self.s3upload_path.exists() and self.md5sum_path.exists():
            return 'all'
        elif self.fastp_path.exists() and self.md5sum_path.exists():
            return 'sequencing_m5sum'
        elif self.fastp_path.exists():
            return 'sequencing'
        elif self.s3upload_path.exists() and self.md5sum_path.exists():
            return 'download_m5sum'
        elif self.s3upload_path.exists():
            return 'download'
        else:
            raise ValueError("No valid input files found")

    def create_all_report(self) -> None: 
        sequencing_info = self.create_sequencing_info(self.fastp_path)
        download_info = self.create_download_info_with_md5sum(self.s3upload_path, self.md5sum_path)
        
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            sequencing_info.to_excel(writer, sheet_name='Sequencing Info', index=False)
            download_info.to_excel(writer, sheet_name='Download Info', index=False)
            
            apply_table_style(writer.sheets['Sequencing Info'], sequencing_info)
            apply_table_style(writer.sheets['Download Info'], download_info)
            
            adjust_column_widths(writer.sheets['Sequencing Info'], sequencing_info)
            adjust_column_widths(writer.sheets['Download Info'], download_info)
            
            format_column_as_number(writer.sheets['Sequencing Info'], sequencing_info, 'Total Reads')
            format_column_as_number(writer.sheets['Sequencing Info'], sequencing_info, 'Total Bases')
            format_column_as_number(writer.sheets['Download Info'], download_info, 'File Size (Bytes)')
            add_hyperlinks(writer.sheets['Download Info'], download_info)
    
    def create_sequencing_md5sum_report(self) -> None:
        sequencing_info = self.create_sequencing_info(self.fastp_path)
        download_info = self.create_md5sum_info(self.md5sum_path)
        
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            sequencing_info.to_excel(writer, sheet_name='Sequencing Info', index=False)
            download_info.to_excel(writer, sheet_name='Download Info', index=False)
            
            apply_table_style(writer.sheets['Sequencing Info'], sequencing_info)
            apply_table_style(writer.sheets['Download Info'], download_info)
            
            adjust_column_widths(writer.sheets['Sequencing Info'], sequencing_info)
            adjust_column_widths(writer.sheets['Download Info'], download_info)
            
            format_column_as_number(writer.sheets['Sequencing Info'], sequencing_info, 'Total Bases')
            format_column_as_number(writer.sheets['Sequencing Info'], sequencing_info, 'Total Reads')
            format_column_as_number(writer.sheets['Download Info'], download_info, 'File Size (Bytes)')
            add_hyperlinks(writer.sheets['Download Info'], download_info)

    def create_sequencing_report(self) -> None:
        sequencing_info = self.create_sequencing_info(self.fastp_path)
        
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            sequencing_info.to_excel(writer, sheet_name='Sequencing Info', index=False)
            apply_table_style(writer.sheets['Sequencing Info'], sequencing_info)
            adjust_column_widths(writer.sheets['Sequencing Info'], sequencing_info)
            format_column_as_number(writer.sheets['Sequencing Info'], sequencing_info, 'Total Reads')
            format_column_as_number(writer.sheets['Sequencing Info'], sequencing_info, 'Total Bases')

    def create_download_md5sum_report(self) -> None:
        download_info = self.create_download_info_with_md5sum(self.s3upload_path, self.md5sum_path)
        
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            download_info.to_excel(writer, sheet_name='Download Info', index=False)
            apply_table_style(writer.sheets['Download Info'], download_info)
            adjust_column_widths(writer.sheets['Download Info'], download_info)
            format_column_as_number(writer.sheets['Download Info'], download_info, 'File Size (Bytes)')
            add_hyperlinks(writer.sheets['Download Info'], download_info)

    def create_download_report(self) -> None:
        download_info = self.create_download_info(self.s3upload_path)
        
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            download_info.to_excel(writer, sheet_name='Download Info', index=False)
            apply_table_style(writer.sheets['Download Info'], download_info)
            adjust_column_widths(writer.sheets['Download Info'], download_info)
            format_column_as_number(writer.sheets['Download Info'], download_info, 'File Size (Bytes)')
            add_hyperlinks(writer.sheets['Download Info'], download_info)






def main(args: argparse.Namespace):

    rg = ReportGenerator(args)

    if rg.report_type in ['all']:
        rg.create_all_report()
    elif rg.report_type in ['sequencing_m5sum']:
        rg.create_sequencing_md5sum_report()
    elif rg.report_type in ['sequencing']:
        rg.create_sequencing_report()
    elif rg.report_type in ['download_m5sum']:
        rg.create_download_md5sum_report()
    elif rg.report_type in ['download']:
        rg.create_download_report()


if __name__ == '__main__':
    args = parse_args()
    main(args) 



